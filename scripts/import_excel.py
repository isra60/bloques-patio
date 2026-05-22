import json
import re
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def slug(value):
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-zA-Z0-9]+", "-", normalized.lower()).strip("-")
    return normalized or "item"


def parse_date(text):
    matches = re.findall(r"\((\d{1,2})[-\s](\d{1,2})[-\s](\d{2,4})\)", text)
    if not matches:
        return None
    day, month, year = matches[-1]
    year = int(year)
    if year < 100:
        year += 2000
    return f"{year:04d}-{int(month):02d}-{int(day):02d}"


def parse_stock(text):
    match = re.search(r"(-?\d+(?:[.,]\d+)?)\s*(?:P\b|PALET|PALETS)", text, re.IGNORECASE)
    if not match:
        match = re.search(r"(-?\d+(?:[.,]\d+)?)\s*\(\d{1,2}[-\s]\d{1,2}[-\s]\d{2,4}\)", text)
    if not match:
        return 0
    return float(match.group(1).replace(",", "."))


def parse_variant_name(text):
    name = re.sub(r"\([^)]*\)", "", text)
    name = re.sub(r"\b-?\d+(?:[.,]\d+)?\s*(?:P\b|PALET|PALETS)\b", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\b-?\d+(?:[.,]\d+)?\s*$", "", name)
    name = re.sub(r"\s+", " ", name).strip(" -")
    return name or text.strip()


def commercial_name(text):
    text_upper = text.upper()
    if "JESUS" in text_upper:
        return "JESUS"
    if "FERNANDO" in text_upper or "FDO" in text_upper:
        return "FERNANDO"
    return text_upper.strip()


def row_has_clients(ws, row):
    return any("CLIENT" in clean(ws.cell(row, col).value).upper() for col in range(1, ws.max_column + 1))


def non_empty_headers(ws, row):
    headers = []
    for col in range(1, ws.max_column + 1):
        value = clean(ws.cell(row, col).value)
        if value:
            headers.append((col, value))
    return headers


def non_empty_rows(ws):
    rows = []
    for row in range(1, ws.max_row + 1):
        if any(clean(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)):
            rows.append(row)
    return rows


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    products = []
    variants = []
    orders = []

    for product_pos, ws in enumerate(wb.worksheets):
        rows = non_empty_rows(ws)
        commercial_row = next((row for row in rows if row_has_clients(ws, row)), None)
        if not commercial_row:
            continue
        previous_rows = [row for row in rows if row < commercial_row]
        if not previous_rows:
            continue
        variant_row = previous_rows[-1]
        title_row = previous_rows[0]
        if title_row != variant_row:
            product_name = next((clean(ws.cell(title_row, col).value) for col in range(1, ws.max_column + 1) if clean(ws.cell(title_row, col).value)), ws.title)
        else:
            product_name = ws.title

        product_id = slug(ws.title)
        products.append({
            "id": product_id,
            "name": product_name,
            "source_sheet": ws.title,
            "position": product_pos,
        })

        header_cells = non_empty_headers(ws, variant_row)
        header_cells = [(col, value) for col, value in header_cells if "CLIENT" not in value.upper()]

        for variant_pos, (start_col, header) in enumerate(header_cells):
            next_cols = [col for col, _ in header_cells if col > start_col]
            end_col = (min(next_cols) - 1) if next_cols else ws.max_column
            variant_id = f"{product_id}-{slug(header)}"
            variants.append({
                "id": variant_id,
                "product_id": product_id,
                "name": parse_variant_name(header),
                "raw_header": header,
                "stock_pallets": parse_stock(header),
                "stock_date": parse_date(header),
                "position": product_pos * 100 + variant_pos,
            })

            commercial_cols = []
            for col in range(start_col, end_col + 1):
                label = clean(ws.cell(commercial_row, col).value)
                if "CLIENT" in label.upper():
                    commercial_cols.append((col, col + 2, commercial_name(label)))

            for name_col, qty_col, commercial in commercial_cols:
                if qty_col > ws.max_column:
                    continue
                for row in range(commercial_row + 1, ws.max_row + 1):
                    customer = clean(ws.cell(row, name_col).value)
                    pallets = ws.cell(row, qty_col).value
                    if not customer and pallets in (None, ""):
                        continue
                    if not customer or pallets in (None, ""):
                        continue
                    try:
                        pallets_number = float(str(pallets).replace(",", "."))
                    except ValueError:
                        continue
                    if pallets_number <= 0:
                        continue
                    orders.append({
                        "variant_id": variant_id,
                        "commercial": commercial,
                        "customer": customer,
                        "pallets": pallets_number,
                        "notes": None,
                        "source_key": f"{variant_id}:{commercial}:{row}:{name_col}",
                    })

    return {
        "source": str(path),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "products": products,
        "variants": variants,
        "orders": orders,
    }


def main():
    if len(sys.argv) != 3:
        print("Usage: import_excel.py INPUT.xlsx OUTPUT.json", file=sys.stderr)
        return 2
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    data = parse_workbook(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"{len(data['products'])} products, {len(data['variants'])} variants, {len(data['orders'])} orders")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
